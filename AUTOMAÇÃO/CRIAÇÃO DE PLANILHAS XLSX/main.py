import openpyxl
import openpyxl.workbook

wb = openpyxl.Workbook() #criando uma planilha vazia
wb.create_sheet('teste') #nome da planilha
wb.remove(wb['Sheet'])

planilha_teste = wb['teste']
planilha_teste.append(['PRODUTO', 'VALOR', 'DATA'])
#planilha_teste['COLUNA + LINHA'] = 'Valor' -> Para adicionar um valor pela coluna e linha

# Lista de valores
valores = ['Maçã', 'Uva', 'Banana']
# Lista de posições
posicoes = ['A2', 'A3', 'A4']

# Adicionando valores às posições especificadas
for valor, posicao in zip(valores, posicoes):
    planilha_teste[posicao] = valor

wb.save('teste.xlsx')
#print(wb.sheetnames)
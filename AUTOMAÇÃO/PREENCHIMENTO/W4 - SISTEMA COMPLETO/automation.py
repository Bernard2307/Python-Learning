from openpyxl import load_workbook
import pyautogui 
import time

planilha = load_workbook('./AUTOMAÇÃO/PREENCHIMENTO/W4 - SISTEMA COMPLETO/vendas_de_produtos.xlsx')
planilha_fornecedores = planilha['vendas']

time.sleep(5)

for linha in planilha_fornecedores.iter_rows(min_row=2,max_row=6,values_only=True):
    fornecedor, produto, quantidade, categoria, pagamento = linha
    #Click and Write para cada coluna
    pyautogui.click(765,332,duration=0.2)
    pyautogui.write(fornecedor)
    pyautogui.click(767,364,duration=0.2)
    pyautogui.write(produto)
    pyautogui.click(766,386,duration=0.2)
    pyautogui.write(str(quantidade))
    pyautogui.click(768,408,duration=0.2)
    pyautogui.write(categoria)
    pyautogui.click(768,438,duration=0.2)
    pyautogui.write(pagamento)
    #Bottons
    pyautogui.click(610,462,duration=1)
    pyautogui.click(661,426,duration=1)

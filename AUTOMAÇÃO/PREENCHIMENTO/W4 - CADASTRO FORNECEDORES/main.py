from openpyxl import load_workbook
import pyautogui

planilha = load_workbook('./AUTOMAÇÃO/PREENCHIMENTO/W4 - CADASTRO FORNECEDORES/vendas_de_produtos.xlsx')
planilha_fornecedores = planilha['vendas']

for linha in planilha_fornecedores.iter_rows(min_row=2,max_row=6,values_only=True):
    cliente, produto, quantidade, categoria = linha 

    pyautogui.click(723,344,duration=0)
    pyautogui.write(cliente)
    pyautogui.click(717,371,duration=0)
    pyautogui.write(produto)
    pyautogui.click(713,402,duration=0)
    pyautogui.write(str(quantidade))
    pyautogui.click(712,423,duration=0)
    pyautogui.write(categoria)
    pyautogui.click(690,455,duration=0) #btn salvar
    pyautogui.click(664,426,duration=1) #btn ok


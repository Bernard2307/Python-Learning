from openpyxl import load_workbook
import pyautogui
import keyboard 

planilha = load_workbook('./W3 - CADASTRO AUTOMATICO/vendas_de_produtos.xlsx')
planilha_produtos = planilha['vendas']

for linha in planilha_produtos.iter_rows(min_row=2,values_only=True):
    cliente, produto, quantidade, categoria = linha #unpacking

    pyautogui.click(657,360,duration=1.5)
    pyautogui.write(cliente)
    pyautogui.click(642,387,duration=1.5)
    pyautogui.write(produto)
    pyautogui.click(665,414,duration=1.5)
    pyautogui.write(str(quantidade))
    pyautogui.click(742,436,duration=1.5)
    pyautogui.write(categoria)
    pyautogui.click(593,463,duration=1.5)
    pyautogui.click(669,426,duration=1.5)

    if keyboard.is_pressed('q'):
        break